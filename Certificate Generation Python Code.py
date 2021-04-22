import cv2 as cv
import openpyxl
import pandas
from PIL import Image
from PIL import Image, ImageDraw, ImageFont
template_path = 'C:/Users/kunch/Desktop/USD/Spring 21/Subjects/Seminar/Assignments/assignment6/Certificate generation code/USDCer.PNG'
details_path = 'C:/Users/kunch/Desktop/USD/Spring 21/Subjects/Seminar/Assignments/assignment6/Certificate generation code/names.xlsx'
output_path = 'C:/Users/kunch/Desktop/USD/Spring 21/Subjects/Seminar/Assignments/assignment6/Certificate generation code/Certificates'
font_scale = 3
font_color = (0,0,0)
coordinate_y_adjustment = 53
coordinate_x_adjustment = 425
obj = openpyxl.load_workbook(details_path)
sheet = obj.active
for i in range(1,10):
    get_name = sheet.cell(row = i ,column = 1)
    first_name=sheet.cell(row = i ,column = 2)
    uni=sheet.cell(row = i ,column = 3)
    certi_name = get_name.value+''+first_name.value
    img = cv.imread(template_path)
    font = cv.FONT_HERSHEY_SCRIPT_SIMPLEX
    text_size = cv.getTextSize(certi_name, font, font_scale, 8)[0] 
    text_x = (img.shape[1] - text_size[0]) / 2+ coordinate_x_adjustment 
    text_y = (img.shape[0] + text_size[1]) / 2 - coordinate_y_adjustment
    text_x = int(text_x)
    text_y = int(text_y)
    cv.putText(img, certi_name,(text_x ,text_y ),font,font_scale,font_color, 4)
    certi_path = output_path + '/certi'+certi_name+ '.png'
    cv.imwrite(certi_path,img)
for i in details_path:
    im = Image.open(r'C:/Users/kunch/Desktop/USD/Spring 21/Subjects/Seminar/Assignments/assignment6/Certificate generation code/USDCer.PNG')
    d = ImageDraw.Draw(im)
    location = (100, 398)
    text_color = (0, 137, 209)
    font = ImageFont.truetype("arial.ttf", 120)
    d.text(location, i, fill = text_color, font = font)
